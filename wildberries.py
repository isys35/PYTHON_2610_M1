import requests
from datetime import datetime
from bs4 import BeautifulSoup
import re
import  openpyxl as ox

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.142 Safari/537.36'
}
wb = ox.Workbook()
ws = wb.worksheets[0]
num = 1
for page in range(1,11):
    URL = "https://catalog.wb.ru/catalog/electronic15/catalog?subject=2872;2875&limit=100&sort=popular&page=" + str(page) + "&appType=128&curr=byn&locale=by&lang=ru&dest=12358386,12358404,3,-59208&regions=1,4,22,30,31,33,40,48,66,68,69,70,80,83&emp=0&reg=1&pricemarginCoeff=1.0&offlineBonus=0&onlineBonus=0&spp=0"
    print(f"Подключаюсь к {URL} ")
    response = requests.get(URL, headers=headers)
    data = response.json()["data"]
    data = data["products"]
    for item in data:
       print(item['name'],item['salePriceU']/100, datetime.now().date())
       ws.cell(column=1, row=num).value = item['name']
       ws.cell(column=2, row=num).value = item['salePriceU']/100
       ws.cell(column=3, row=num).value = datetime.now().date()
       num = num + 1
wb.save('some1.xlsx')


