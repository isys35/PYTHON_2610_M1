import requests
import json
from openpyxl import Workbook
from datetime import datetime

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/109.0.0.0 Safari/537.36'}


def parse_computers():
    wb = Workbook()
    ws = wb.worksheets[0]
    num = 1
    for page in range(1, 11):
        url = 'https://catalog.wb.ru/catalog/electronic15/catalog?cat=9963&limit=100&sort=popular&page=' + str(page) + \
              '&xsubject=2875&appType=128&curr=byn&locale=by&lang=ru&dest=12358386,12358404,3,-59208&' \
              'regions=1,4,22,30,31,33,40,48,66,68,69,70,80,83&emp=0&reg=1&pricemarginCoeff=1.0&' \
              'offlineBonus=0&onlineBonus=0&spp=0'
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()

        if response.status_code >= 200 and response.status_code < 400:
            with open("parse_computer.json", "w") as json_file:
                json.dump(response.json(), json_file, indent=4)
            for item in response.json()["data"]["products"]:
                print(f'{item["name"]}, Цена со скидкой: {item["salePriceU"] / 100}, Дата и время: {datetime.now()}')
                ws.cell(column=1, row=num).value = item['name']
                ws.cell(column=2, row=num).value = item['salePriceU'] / 100
                ws.cell(column=3, row=num).value = datetime.now()
                num = num + 1
            wb.save('parser.xlsx')
        else:
            print('Not Found.', 'code status = ', response.status_code)


if __name__ == '__main__':
    parse_computers()
